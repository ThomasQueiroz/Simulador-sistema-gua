import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import random
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class LoginDialog(simpledialog.Dialog):
    def __init__(self, parent, title=None):  # CORREÇÃO: __init__ em vez de _init_
        self.username = None
        self.password = None
        super().__init__(parent, title)

    def body(self, master):
        ttk.Label(master, text="Usuário:", font=("Arial", 10)).grid(row=0, sticky=tk.W, pady=5)
        ttk.Label(master, text="Senha:", font=("Arial", 10)).grid(row=1, sticky=tk.W, pady=5)

        self.user_entry = ttk.Entry(master, width=20)
        self.pass_entry = ttk.Entry(master, width=20, show="*")

        self.user_entry.grid(row=0, column=1, padx=5, pady=5)
        self.pass_entry.grid(row=1, column=1, padx=5, pady=5)

        return self.user_entry

    def apply(self):
        self.username = self.user_entry.get()
        self.password = self.pass_entry.get()


class WaterMeterSimulator:
    def __init__(self, root):  # CORREÇÃO: __init__ em vez de _init_
        self.root = root
        self.root.title("Simulador de Medidor de Água")
        self.root.geometry("900x700")
        self.root.resizable(True, True)

        # Configurar estilo com tema verde
        self.setup_styles()

        # Inicializar variáveis
        self.current_liters = 0
        self.cycle = 1
        # targets agora é variável de sessão (modificável pelo usuário)
        self.targets = [1000, 1500, 2000, 2500, 3000]
        self.current_target_index = 0
        self.data = []
        self.filename = "registro_agua.csv"
        self.current_user = None
        self.password = "agua123"  # Senha fixa para acesso
        # Variáveis de controle do auto-mode (precisa vir antes do define_metrics!)
        self.auto_mode = False
        self.auto_job = None

        # Verificar login
        if not self.do_login():
            self.root.destroy()
            return
        self.setup_ui()

        # Depois do login, pedir ao usuário para definir as métricas (opcional)
        self.define_metrics(initial=True)


        # Carregar dados existentes se o arquivo já existir
        if os.path.exists(self.filename):
            try:
                # Ler com BOM UTF-8 para evitar problemas de acentuação no Excel
                self.df = pd.read_csv(self.filename, encoding='utf-8-sig')
            except Exception:
                # Se houver erro ao ler, recriar DataFrame padrão
                self.df = pd.DataFrame(columns=["Ciclo", "Marca (L)", "Horário", "Usuário"])
        else:
            self.df = pd.DataFrame(columns=["Ciclo", "Marca (L)", "Horário", "Usuário"])

        # Iniciar simulação
        self.update_display()
        self.update_treeview()

    def do_login(self):
        """Realiza o login do usuário"""
        login_attempts = 0
        max_attempts = 3

        while login_attempts < max_attempts:
            login_dialog = LoginDialog(self.root, "Login do Sistema")

            if login_dialog.username is None:  # Usuário cancelou
                return False

            if login_dialog.password == self.password:
                self.current_user = login_dialog.username
                return True
            else:
                login_attempts += 1
                remaining = max_attempts - login_attempts
                messagebox.showerror("Erro de Login",
                                     f"Senha incorreta! Tentativas restantes: {remaining}")

        messagebox.showerror("Erro de Login", "Número máximo de tentativas excedido!")
        return False

    def setup_styles(self):
        style = ttk.Style()
        style.theme_use('clam')  # Usar um tema que permite customização

        # Configurar cores
        self.bg_color = "#e8f5e9"  # Verde muito claro
        self.accent_color = "#388e3c"  # Verde escuro
        self.light_accent = "#c8e6c9"  # Verde claro
        self.text_color = "#1b5e20"  # Verde texto escuro

        # Configurar estilo dos frames
        style.configure("Main.TFrame", background=self.bg_color)
        style.configure("Button.TFrame", background=self.bg_color)
        style.configure("Status.TFrame", background=self.light_accent)

        # Configurar estilo dos labels
        style.configure("Title.TLabel", font=("Arial", 16, "bold"), foreground=self.text_color,
                        background=self.bg_color)
        style.configure("Subtitle.TLabel", font=("Arial", 12, "bold"), foreground=self.text_color,
                        background=self.bg_color)
        style.configure("Regular.TLabel", font=("Arial", 10), foreground=self.text_color, background=self.bg_color)
        style.configure("Status.TLabel", font=("Arial", 10), foreground=self.text_color, background=self.light_accent)

        # Configurar estilo dos botões
        style.configure("Accent.TButton", font=("Arial", 10, "bold"),
                        foreground="white", background=self.accent_color,
                        focuscolor=style.configure(".")["background"])
        style.map("Accent.TButton",
                  background=[('active', '#2e7d32')],  # Verde mais escuro quando ativo
                  foreground=[('active', 'white')])

        # Configurar estilo da barra de progresso
        style.configure("Green.Horizontal.TProgressbar",
                        background=self.accent_color,
                        troughcolor=self.light_accent)

        # Configurar estilo da treeview
        style.configure("Treeview",
                        fieldbackground=self.bg_color,
                        background="white",
                        foreground=self.text_color,
                        rowheight=25)
        style.configure("Treeview.Heading",
                        font=("Arial", 10, "bold"),
                        background=self.accent_color,
                        foreground="white")

        # Configurar cor de fundo da janela principal
        self.root.configure(background=self.bg_color)

    def setup_ui(self):
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="15", style="Main.TFrame")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Configurar pesos para redimensionamento
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(7, weight=1)  # Dar mais espaço para a treeview

        # Título com informações do usuário
        title_text = f"💧 Simulador de Medidor de Água - Usuário: {self.current_user}"
        title_label = ttk.Label(main_frame, text=title_text, style="Title.TLabel")
        title_label.grid(row=0, column=0, columnspan=4, pady=(0, 10))

        # Informações atuais
        ttk.Label(main_frame, text="Litros atuais:", style="Subtitle.TLabel").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.liters_label = ttk.Label(main_frame, text="0 L", style="Subtitle.TLabel")
        self.liters_label.grid(row=1, column=1, sticky=tk.W, pady=5)

        ttk.Label(main_frame, text="Próxima marca:", style="Subtitle.TLabel").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.target_label = ttk.Label(main_frame, text="1000 L", style="Subtitle.TLabel")
        self.target_label.grid(row=2, column=1, sticky=tk.W, pady=5)

        ttk.Label(main_frame, text="Ciclo atual:", style="Subtitle.TLabel").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.cycle_label = ttk.Label(main_frame, text="1", style="Subtitle.TLabel")
        self.cycle_label.grid(row=3, column=1, sticky=tk.W, pady=5)

        # Barra de progresso
        ttk.Label(main_frame, text="Progresso:", style="Subtitle.TLabel").grid(row=4, column=0, sticky=tk.W,
                                                                               pady=(20, 5))
        self.progress = ttk.Progressbar(main_frame, orient=tk.HORIZONTAL, length=400,
                                        mode='determinate', style="Green.Horizontal.TProgressbar")
        self.progress.grid(row=4, column=1, columnspan=3, sticky=(tk.W, tk.E), pady=(20, 5))

        # Botões de controle
        button_frame = ttk.Frame(main_frame, style="Button.TFrame")
        button_frame.grid(row=5, column=0, columnspan=4, pady=20)

        self.add_water_btn = ttk.Button(button_frame, text="Adicionar 100L", command=lambda: self.add_water(100),
                                        style="Accent.TButton")
        self.add_water_btn.pack(side=tk.LEFT, padx=5)

        self.add_custom_btn = ttk.Button(button_frame, text="Definir Quantidade", command=self.set_custom_amount,
                                         style="Accent.TButton")
        self.add_custom_btn.pack(side=tk.LEFT, padx=5)

        # Novo botão: Registrar Valor Manual
        self.manual_btn = ttk.Button(button_frame, text="Registrar Valor Manual", command=self.register_manual_value,
                                     style="Accent.TButton")
        self.manual_btn.pack(side=tk.LEFT, padx=5)

        # Novo botão: Definir Métricas
        self.define_targets_btn = ttk.Button(button_frame, text="Definir Métricas", command=self.define_metrics,
                                            style="Accent.TButton")
        self.define_targets_btn.pack(side=tk.LEFT, padx=5)

        self.auto_btn = ttk.Button(button_frame, text="Modo Automático", command=self.toggle_auto_mode,
                                   style="Accent.TButton")
        self.auto_btn.pack(side=tk.LEFT, padx=5)

        self.reset_btn = ttk.Button(button_frame, text="Reiniciar Sistema", command=self.reset_system,
                                    style="Accent.TButton")
        self.reset_btn.pack(side=tk.LEFT, padx=5)

        self.delete_btn = ttk.Button(button_frame, text="Deletar Tudo", command=self.delete_all, style="Accent.TButton")
        self.delete_btn.pack(side=tk.LEFT, padx=5)

        self.export_btn = ttk.Button(button_frame, text="Baixar Planilha", command=self.export_spreadsheet,
                                     style="Accent.TButton")
        self.export_btn.pack(side=tk.LEFT, padx=5)

        # Visualização da planilha
        ttk.Label(main_frame, text="Registros na Planilha:", style="Subtitle.TLabel").grid(row=6, column=0,
                                                                                           columnspan=4, sticky=tk.W,
                                                                                           pady=(20, 5))

        # Treeview para mostrar os dados
        columns = ("Ciclo", "Marca (L)", "Horário", "Usuário")
        self.tree = ttk.Treeview(main_frame, columns=columns, show="headings", height=12)

        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=120, anchor=tk.CENTER)

        self.tree.grid(row=7, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))

        # Adicionar scrollbar
        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.grid(row=7, column=3, sticky=(tk.N, tk.S))

        # Frame de status
        status_frame = ttk.Frame(main_frame, style="Status.TFrame", padding="5")
        status_frame.grid(row=8, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=(10, 0))

        ttk.Label(status_frame, text="Status:", style="Status.TLabel").pack(side=tk.LEFT)
        self.status_label = ttk.Label(status_frame, text="Pronto para simular", style="Status.TLabel")
        self.status_label.pack(side=tk.LEFT, padx=5)

        # Configurar auto-mode
        self.auto_mode = False
        self.auto_job = None

    def add_water(self, amount):
        """Adiciona a quantidade especificada de água"""
        self.current_liters += amount
        self.update_display()

        # Verificar se atingiu a marca atual
        if self.current_target_index < len(self.targets) and self.current_liters >= self.targets[self.current_target_index]:
            self.record_marker()

            # Avançar para a próxima marca ou reiniciar ciclo
            self.current_target_index += 1

            if self.current_target_index >= len(self.targets):
                self.cycle += 1
                self.current_target_index = 0
                self.current_liters = 0
                self.status_label.config(text="Ciclo completo! Reiniciando...")

            self.update_display()

    def set_custom_amount(self):
        """Define uma quantidade específica de água (substitui o valor atual)"""
        # Parar o modo automático se estiver ativo
        if self.auto_mode:
            self.toggle_auto_mode()

        # Solicitar a quantidade ao usuário
        amount_str = simpledialog.askstring("Definir Quantidade",
                                            "Digite o valor de litros:",
                                            parent=self.root)

        if amount_str:
            try:
                new_amount = int(amount_str)
                if new_amount < self.current_liters:
                    messagebox.showerror("Erro", "O novo valor não pode ser menor que o atual!")
                    return

                # Substituir o valor atual
                self.current_liters = new_amount
                self.update_display()

                # Verificar se atingiu alguma marca
                while self.current_target_index < len(self.targets) and self.current_liters >= self.targets[self.current_target_index]:
                    self.record_marker()
                    self.current_target_index += 1

                    if self.current_target_index >= len(self.targets):
                        self.cycle += 1
                        self.current_target_index = 0
                        self.current_liters = 0
                        self.status_label.config(text="Ciclo completo! Reiniciando...")
                        break

                self.update_display()
                self.status_label.config(text=f"Valor definido para {new_amount}L")

            except ValueError:
                messagebox.showerror("Erro", "Por favor, digite um número válido!")

    def get_last_recorded_value(self):
        """Retorna o último valor registrado na coluna 'Marca (L)' (ou 0 se não houver registros).

        Antes essa função retornava o valor máximo encontrado, o que fazia com que um registro antigo
        alto bloqueasse registros válidos na sessão atual. Agora retornamos o último valor inserido.
        """
        if getattr(self, 'df', None) is None or self.df.empty:
            return 0
        try:
            # Converter para numérico e pegar o último valor não-nulo (na ordem do DataFrame)
            numeric = pd.to_numeric(self.df['Marca (L)'], errors='coerce')
            non_null = numeric.dropna()
            if non_null.empty:
                return 0
            # Pegar o último valor (última linha do DataFrame)
            return int(non_null.iloc[-1])
        except Exception:
            return 0

    def register_manual_value(self):
        """Registrar um valor manualmente (adiciona à tabela e atualiza current_liters).

        Valida que o valor informado não seja menor que o maior já registrado e nem menor que o current_liters.
        """
        # Parar o modo automático se estiver ativo
        if self.auto_mode:
            self.toggle_auto_mode()

        amount_str = simpledialog.askstring("Registrar Valor Manual",
                                            "Digite o valor de litros a registrar:",
                                            parent=self.root)
        if not amount_str:
            return

        try:
            value = int(amount_str)
        except ValueError:
            messagebox.showerror("Erro", "Por favor, digite um número válido!")
            return

        last_recorded = self.get_last_recorded_value()
        # Impedir se menor que último registrado ou menor que current_liters
        if value < last_recorded or value < self.current_liters:
            messagebox.showerror("Erro", f"O valor deve ser maior ou igual ao último registrado ({last_recorded}L) e ao valor atual ({self.current_liters}L).")
            return

        # Atualizar current_liters
        self.current_liters = value
        self.update_display()

        # Adicionar registro manual à DataFrame
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        new_data = {
            "Ciclo": self.cycle,
            "Marca (L)": value,
            "Horário": timestamp,
            "Usuário": self.current_user
        }

        new_row = pd.DataFrame([new_data])
        self.df = pd.concat([self.df, new_row], ignore_index=True)

        # Garantir colunas
        expected_columns = ["Ciclo", "Marca (L)", "Horário", "Usuário"]
        for col in expected_columns:
            if col not in self.df.columns:
                self.df[col] = None

        # Salvar e atualizar
        self.df.to_csv(self.filename, index=False, encoding='utf-8-sig')
        self.update_treeview()
        self.status_label.config(text=f"Registro manual de {value}L adicionado às {timestamp}")

    def validate_manual_input(self, value):
        """Valida se o valor digitado é >= ao atual da sessão"""

        if value < self.current_liters:
            messagebox.showerror(
                "Erro",
                f"O valor deve ser maior ou igual ao valor atual da sessão ({self.current_liters} L)."
            )
            return False
        return True


    def record_marker(self):
        # Registrar a marca atual na planilha
        # Proteção se índice estiver fora de alcance
        if self.current_target_index >= len(self.targets):
            return

        marker = self.targets[self.current_target_index]
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        new_data = {
            "Ciclo": self.cycle,
            "Marca (L)": marker,
            "Horário": timestamp,
            "Usuário": self.current_user
        }

        self.data.append(new_data)

        # Adicionar ao DataFrame
        new_row = pd.DataFrame([new_data])
        self.df = pd.concat([self.df, new_row], ignore_index=True)

        # Garantir que todas as colunas existem
        expected_columns = ["Ciclo", "Marca (L)", "Horário", "Usuário"]
        for col in expected_columns:
            if col not in self.df.columns:
                self.df[col] = None

        # Salvar no arquivo CSV
        self.df.to_csv(self.filename, index=False, encoding='utf-8-sig')

        # Atualizar a exibição
        self.update_treeview()

        # Mostrar mensagem de status
        self.status_label.config(text=f"Marca de {marker}L registrada às {timestamp}")

    def update_display(self):
        # Atualizar labels
        self.liters_label.config(text=f"{self.current_liters} L")

        if self.current_target_index < len(self.targets):
            next_target = self.targets[self.current_target_index]
            self.target_label.config(text=f"{next_target} L")

            # Atualizar barra de progresso
            try:
                progress_value = min(100, (self.current_liters / next_target) * 100) if next_target > 0 else 0
            except Exception:
                progress_value = 0
            self.progress['value'] = progress_value
        else:
            self.target_label.config(text="N/A")
            self.progress['value'] = 0

        self.cycle_label.config(text=str(self.cycle))

    def update_treeview(self):
        # Limpar a treeview
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Adicionar os dados mais recentes (todos os registros)
        for _, row in self.df.iterrows():
            self.tree.insert("", "end", values=(
                row.get("Ciclo", ""),
                row.get("Marca (L)", ""),
                row.get("Horário", ""),
                row.get("Usuário", "")
            ))

    def toggle_auto_mode(self):
        self.auto_mode = not self.auto_mode

        if self.auto_mode:
            self.auto_btn.config(text="Parar Modo Automático")
            self.add_water_btn.config(state="disabled")
            self.add_custom_btn.config(state="disabled")
            self.manual_btn.config(state="disabled")
            self.define_targets_btn.config(state="disabled")
            self.status_label.config(text="Modo automático ativado")
            self.run_auto_mode()
        else:
            self.auto_btn.config(text="Modo Automático")
            self.add_water_btn.config(state="normal")
            self.add_custom_btn.config(state="normal")
            self.manual_btn.config(state="normal")
            self.define_targets_btn.config(state="normal")
            self.status_label.config(text="Modo automático desativado")
            if self.auto_job:
                self.root.after_cancel(self.auto_job)

    def run_auto_mode(self):
        if self.auto_mode:
            # Adicionar uma quantidade aleatória de água entre 50 e 200 litros
            amount = random.randint(50, 200)
            self.add_water(amount)

            # Agendar próxima execução
            delay = random.randint(500, 2000)  # Entre 0.5 e 2 segundos
            self.auto_job = self.root.after(delay, self.run_auto_mode)

    def reset_system(self):
        # Reiniciar todas as variáveis
        self.current_liters = 0
        self.cycle = 1
        self.current_target_index = 0
        self.data = []

        # Manter o arquivo CSV existente, mas recomeçar a simulação
        self.update_display()
        self.status_label.config(text="Sistema reiniciado - os dados no arquivo CSV foram mantidos")

    def delete_all(self):
        # Confirmar se o usuário realmente quer deletar tudo
        result = messagebox.askyesno(
            "Confirmar Exclusão",
            "Tem certeza que deseja deletar TODOS os registros?\nEsta ação não pode ser desfeita."
        )

        if result:
            # Deletar o arquivo CSV
            if os.path.exists(self.filename):
                os.remove(self.filename)

            # Reiniciar o DataFrame
            self.df = pd.DataFrame(columns=["Ciclo", "Marca (L)", "Horário", "Usuário"])

            # Reiniciar as variáveis
            self.current_liters = 0
            self.cycle = 1
            self.current_target_index = 0
            self.data = []

            # Atualizar a interface
            self.update_display()
            self.update_treeview()
            self.status_label.config(text="Todos os registros foram deletados")

    def format_excel_sheet(self, ws):
        """Formata a planilha Excel para melhor aparência"""
        # Definir estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        data_font = Font(size=11)
        header_fill = PatternFill(start_color="388e3c", end_color="388e3c", fill_type="solid")
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        center_aligned = Alignment(horizontal="center", vertical="center")
        left_aligned = Alignment(horizontal="left", vertical="center")

        # Ajustar largura das colunas
        column_widths = {
            'A': 10,  # Ciclo
            'B': 15,  # Marca (L)
            'C': 25,  # Horário
            'D': 20  # Usuário
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Formatar cabeçalhos
        for col in range(1, 5):  # Temos 4 colunas
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_aligned

        # Formatar células de dados
        for row in range(2, len(self.df) + 2):
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.font = data_font

                if col in [1, 2]:  # Colunas numéricas
                    cell.alignment = center_aligned
                else:  # Colunas de texto
                    cell.alignment = left_aligned

        # Congelar painel superior (cabeçalho)
        ws.freeze_panes = "A2"

        # Adicionar filtros aos cabeçalhos
        ws.auto_filter.ref = f"A1:D{len(self.df) + 1}"

    def export_spreadsheet(self):
        # Verificar if há dados para exportar
        if self.df.empty:
            messagebox.showwarning("Aviso", "Não há dados para exportar!")
            return

        # Abrir diálogo para escolher onde salvar o arquivo
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("CSV Files", "*.csv"), ("All Files", "*.*")],  # CORREÇÃO AQUI
            title="Salvar Planilha Como"
        )

        if file_path:  # Se o usuário não cancelou
            try:
                if file_path.endswith('.xlsx'):
                    # Criar um novo workbook e worksheet
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Registro de Água"

                    # Adicionar cabeçalhos
                    headers = list(self.df.columns)
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_idx, value=header)

                    # Adicionar dados - garantir que cada valor vai para sua própria célula
                    for row_idx, (_, row_data) in enumerate(self.df.iterrows(), 2):
                        ws.cell(row=row_idx, column=1, value=row_data.get("Ciclo"))
                        ws.cell(row=row_idx, column=2, value=row_data.get("Marca (L)"))
                        ws.cell(row=row_idx, column=3, value=row_data.get("Horário"))
                        ws.cell(row=row_idx, column=4, value=row_data.get("Usuário"))

                    # Aplicar formatação
                    self.format_excel_sheet(ws)

                    # Salvar arquivo
                    wb.save(file_path)
                else:
                    # Exportar para CSV (padrão)
                    self.df.to_csv(file_path, index=False, encoding='utf-8', sep=';')

                messagebox.showinfo("Sucesso", f"Planilha exportada com sucesso para:\n{file_path}")
                self.status_label.config(text=f"Planilha exportada: {os.path.basename(file_path)}")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao exportar planilha:\n{str(e)}")
                self.status_label.config(text="Erro ao exportar planilha")

    def define_metrics(self, initial=False):
        """Permite ao usuário definir métricas (targets) para a sessão atual.

        Se initial=True, mostra o diálogo imediatamente após login (opcional).
        """
        # Se estiver em auto-mode, impedir alteração (ou parar automaticamente?) — vamos parar automaticamente
        if self.auto_mode:
            self.toggle_auto_mode()

        default = ",".join(str(x) for x in self.targets)
        prompt = "Digite as métricas separadas por vírgula (ex: 1000,1500,2000):"
        if initial:
            prompt = f"Defina as métricas para esta sessão (pressione OK para manter os padrões):\nPadrão atual: {default}"

        input_str = simpledialog.askstring("Definir Métricas", prompt, initialvalue=default, parent=self.root)
        if input_str is None:
            return

        # Tentar converter para lista de inteiros
        try:
            values = [int(x.strip()) for x in input_str.split(',') if x.strip() != ""]
            if not values:
                messagebox.showerror("Erro", "Você deve fornecer ao menos uma métrica válida.")
                return
            # Garantir que esteja em ordem crescente e sem duplicatas
            values = sorted(list(set(values)))
            self.targets = values

            # Atualizar current_target_index com base no valor atual de litros
            self.current_target_index = 0
            while self.current_target_index < len(self.targets) and self.current_liters >= self.targets[self.current_target_index]:
                self.current_target_index += 1

            self.update_display()
            self.status_label.config(text=f"Métricas definidas: {', '.join(str(x) for x in self.targets)}")
        except ValueError:
            messagebox.showerror("Erro", "Formato inválido. Use apenas números separados por vírgula.")


if __name__ == "__main__":  # CORREÇÃO: __main__ em vez de _main_
    root = tk.Tk()
    app = WaterMeterSimulator(root)
    root.mainloop()